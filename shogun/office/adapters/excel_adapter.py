"""Excel Adapter — Hybrid openpyxl + COM automation for Excel workbooks.

Uses openpyxl for all read/write operations (cross-platform, no Office needed).
Uses COM (via com_thread_pool) only for:
  - Calculation refresh (Application.Calculate)
  - PDF export (ExportAsFixedFormat)

All operations:
  1. Validate file paths against approved boundaries
  2. Check posture permissions
  3. Acquire the Excel mutex
  4. Execute the operation
  5. Log via EventLogger
  6. Return structured results
"""

from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Any

log = logging.getLogger("shogun.office.adapters.excel")


# ── Workbook Handle ──────────────────────────────────────────────────


class ExcelWorkbookHandle:
    """Tracks an open Excel workbook (openpyxl-based)."""

    def __init__(self, path: Path, workbook: Any):
        self.path = path
        self.workbook = workbook
        self.opened_at = time.time()

    def close(self):
        try:
            self.workbook.close()
        except Exception:
            pass


# ── Core Adapter Functions ───────────────────────────────────────────


def open_workbook(file_path: str) -> ExcelWorkbookHandle:
    """Open an Excel workbook with openpyxl.

    Args:
        file_path: Validated absolute path to the .xlsx file.

    Returns:
        ExcelWorkbookHandle with the loaded workbook.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel operations. Install with: pip install openpyxl")

    path = Path(file_path)

    # Detect password-protected files
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
    except Exception as exc:
        error_str = str(exc).lower()
        if "password" in error_str or "encrypted" in error_str:
            from shogun.office.exceptions import PasswordProtectedError
            raise PasswordProtectedError(str(path))
        if "corrupt" in error_str or "invalid" in error_str or "not a zip" in error_str:
            from shogun.office.exceptions import CorruptedFileError
            raise CorruptedFileError(str(path), str(exc))
        raise

    log.debug("Opened workbook: %s (%d sheets)", path, len(wb.sheetnames))
    return ExcelWorkbookHandle(path=path, workbook=wb)


def close_workbook(handle: ExcelWorkbookHandle) -> None:
    """Close an Excel workbook."""
    handle.close()
    log.debug("Closed workbook: %s", handle.path)


def list_sheets(handle: ExcelWorkbookHandle) -> list[str]:
    """List all sheet names in the workbook."""
    return handle.workbook.sheetnames


def read_used_range(handle: ExcelWorkbookHandle, sheet_name: str) -> list[list[Any]]:
    """Read all used cells from a sheet.

    Returns a 2D list of cell values.
    """
    wb = handle.workbook
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    log.debug("Read used range from '%s': %d rows", sheet_name, len(data))
    return data


def read_range(
    handle: ExcelWorkbookHandle,
    sheet_name: str,
    range_str: str,
) -> list[list[Any]]:
    """Read a specific cell range (e.g. 'B2:D10').

    Returns a 2D list of cell values.
    """
    wb = handle.workbook
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    data = []
    try:
        for row in ws[range_str]:
            data.append([cell.value for cell in row])
    except Exception as exc:
        raise ValueError(f"Invalid range '{range_str}': {exc}")

    log.debug("Read range '%s' from '%s': %d rows", range_str, sheet_name, len(data))
    return data


def read_named_range(handle: ExcelWorkbookHandle, name: str) -> list[list[Any]]:
    """Read a named range from the workbook."""
    wb = handle.workbook
    if name not in wb.defined_names:
        available = list(wb.defined_names.definedName) if hasattr(wb.defined_names, 'definedName') else []
        raise ValueError(f"Named range '{name}' not found. Available: {[d.name for d in available]}")

    defn = wb.defined_names[name]
    dests = defn.destinations
    data = []
    for sheet_title, coord in dests:
        ws = wb[sheet_title]
        for row in ws[coord]:
            if hasattr(row, '__iter__'):
                data.append([cell.value for cell in row])
            else:
                data.append([row.value])

    return data


def write_range(
    handle: ExcelWorkbookHandle,
    sheet_name: str,
    range_str: str,
    values: list[list[Any]],
) -> None:
    """Write values to a specific range.

    Args:
        handle: The workbook handle.
        sheet_name: Target sheet name.
        range_str: Start cell or range (e.g. 'B4' or 'B4:D12').
        values: 2D list of values to write.
    """
    wb = handle.workbook
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist. Available sheets: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Parse the start cell from range
    from openpyxl.utils import range_boundaries
    try:
        # If it's a range like B4:D12, use the top-left
        min_col, min_row, _, _ = range_boundaries(range_str)
    except Exception:
        # If it's a single cell like B4
        from openpyxl.utils import coordinate_to_tuple
        min_row, min_col = coordinate_to_tuple(range_str)

    for row_idx, row_data in enumerate(values):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=min_row + row_idx, column=min_col + col_idx, value=value)

    log.debug("Wrote %d rows to '%s!%s'", len(values), sheet_name, range_str)


def create_sheet(handle: ExcelWorkbookHandle, name: str) -> None:
    """Create a new sheet in the workbook."""
    wb = handle.workbook
    if name in wb.sheetnames:
        raise ValueError(f"Sheet '{name}' already exists.")
    wb.create_sheet(title=name)
    log.debug("Created sheet: %s", name)


def copy_sheet(handle: ExcelWorkbookHandle, source_name: str, target_name: str) -> None:
    """Copy a sheet within the workbook."""
    wb = handle.workbook
    if source_name not in wb.sheetnames:
        raise ValueError(f"Source sheet '{source_name}' does not exist.")
    if target_name in wb.sheetnames:
        raise ValueError(f"Target sheet '{target_name}' already exists.")

    source = wb[source_name]
    target = wb.copy_worksheet(source)
    target.title = target_name
    log.debug("Copied sheet '%s' → '%s'", source_name, target_name)


def apply_basic_formatting(
    handle: ExcelWorkbookHandle,
    sheet_name: str,
    range_str: str,
    formatting: dict[str, Any],
) -> None:
    """Apply basic formatting to a range.

    Supported formatting keys:
      - bold: bool
      - italic: bool
      - font_size: int
      - font_color: str (hex, e.g. 'FF0000')
      - fill_color: str (hex)
      - number_format: str (e.g. '#,##0.00')
    """
    from openpyxl.styles import Font, PatternFill

    wb = handle.workbook
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist.")

    ws = wb[sheet_name]
    font_kwargs: dict[str, Any] = {}
    if "bold" in formatting:
        font_kwargs["bold"] = formatting["bold"]
    if "italic" in formatting:
        font_kwargs["italic"] = formatting["italic"]
    if "font_size" in formatting:
        font_kwargs["size"] = formatting["font_size"]
    if "font_color" in formatting:
        font_kwargs["color"] = formatting["font_color"]

    fill = None
    if "fill_color" in formatting:
        fill = PatternFill(start_color=formatting["fill_color"], end_color=formatting["fill_color"], fill_type="solid")

    number_format = formatting.get("number_format")

    for row in ws[range_str]:
        for cell in row:
            if font_kwargs:
                cell.font = Font(**font_kwargs)
            if fill:
                cell.fill = fill
            if number_format:
                cell.number_format = number_format

    log.debug("Applied formatting to '%s!%s'", sheet_name, range_str)


def save_as(handle: ExcelWorkbookHandle, output_path: str) -> str:
    """Save the workbook to a new file path.

    Args:
        output_path: Validated absolute path for the output file.

    Returns:
        The output file path as a string.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    handle.workbook.save(str(out))
    log.info("Saved workbook to: %s", out)
    return str(out)


def get_workbook_metadata(handle: ExcelWorkbookHandle) -> dict[str, Any]:
    """Get metadata about the workbook."""
    wb = handle.workbook
    props = wb.properties
    return {
        "file": str(handle.path),
        "sheets": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "title": props.title or "",
        "creator": props.creator or "",
        "created": str(props.created) if props.created else "",
        "modified": str(props.modified) if props.modified else "",
        "last_modified_by": props.lastModifiedBy or "",
    }


# ── COM-Only Functions (PDF Export, Calculate) ───────────────────────


def _com_calculate(file_path: str, visible: bool = False) -> None:
    """Refresh all calculations in a workbook via COM.

    Must be called on the STA thread pool.
    """
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = visible
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(Path(file_path).resolve()))
        try:
            wb.Application.Calculate()
            wb.Save()
        finally:
            wb.Close(SaveChanges=True)
    finally:
        excel.Quit()


def _com_export_pdf(file_path: str, output_path: str, visible: bool = False) -> str:
    """Export a workbook to PDF via COM.

    Must be called on the STA thread pool.
    """
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = visible
    excel.DisplayAlerts = False
    out = str(Path(output_path).resolve())
    try:
        wb = excel.Workbooks.Open(str(Path(file_path).resolve()))
        try:
            # xlTypePDF = 0
            wb.ExportAsFixedFormat(0, out)
        finally:
            wb.Close(SaveChanges=False)
    finally:
        excel.Quit()

    log.info("Exported workbook to PDF: %s", out)
    return out


async def calculate(file_path: str, visible: bool = False) -> None:
    """Refresh calculations — async wrapper around COM call."""
    from shogun.office.com_thread_pool import run_com, office_lock
    async with office_lock("excel"):
        await run_com(_com_calculate, file_path, visible)


async def export_pdf(file_path: str, output_path: str, visible: bool = False) -> str:
    """Export to PDF — async wrapper around COM call."""
    from shogun.office.com_thread_pool import run_com, office_lock
    async with office_lock("excel"):
        return await run_com(_com_export_pdf, file_path, output_path, visible)
